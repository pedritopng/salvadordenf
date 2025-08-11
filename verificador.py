import pandas as pd
import tkinter as tk
from tkinter import filedialog
import os
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill


def selecionar_arquivo(titulo):
    """Abre uma janela para o usuário selecionar a planilha de notas."""
    root = tk.Tk()
    root.withdraw()
    caminho = filedialog.askopenfilename(
        title=titulo,
        filetypes=(("Arquivos de Planilha", "*.csv *.xlsx"), ("Todos os arquivos", "*.*"))
    )
    return caminho


def selecionar_pasta(titulo):
    """Abre uma janela para o usuário selecionar a pasta com os PDFs."""
    root = tk.Tk()
    root.withdraw()
    caminho = filedialog.askdirectory(title=titulo)
    return caminho


def carregar_notas_da_planilha(caminho_arquivo):
    """Lê a planilha e retorna um DataFrame e uma lista de notas únicas."""
    print(f"Lendo a planilha: {os.path.basename(caminho_arquivo)}...")
    try:
        if caminho_arquivo.endswith('.csv'):
            df = pd.read_csv(caminho_arquivo, encoding='latin-1', delimiter=';')
        else:
            df = pd.read_excel(caminho_arquivo)

        df.columns = df.columns.str.strip()

        if 'DOCUMENTO' not in df.columns:
            print("ERRO: A coluna 'DOCUMENTO' não foi encontrada na planilha.")
            return None, None

        # Pega todas as notas, mantendo a ordem original para o relatório
        notas_da_planilha = [str(int(nota)) for nota in df['DOCUMENTO'] if pd.notna(nota)]
        # Pega apenas as notas únicas para a verificação
        notas_unicas_planilha = set(notas_da_planilha)

        print(f"{len(notas_unicas_planilha)} notas fiscais únicas encontradas na planilha.")
        return df, notas_unicas_planilha
    except Exception as e:
        print(f"ERRO ao ler a planilha: {e}")
        return None, None


def escanear_pasta_pdf(caminho_pasta):
    """Escaneia a pasta, extrai os números das notas dos nomes dos arquivos PDF e retorna um conjunto."""
    print(f"Escaneando a pasta de PDFs: {caminho_pasta}...")
    notas_encontradas = set()
    try:
        for nome_arquivo in os.listdir(caminho_pasta):
            if nome_arquivo.lower().endswith('.pdf'):
                # Usa expressão regular para encontrar o número no nome do arquivo
                match = re.search(r'(\d+)', nome_arquivo)
                if match:
                    notas_encontradas.add(match.group(0))

        print(f"{len(notas_encontradas)} arquivos PDF com números de nota foram encontrados.")
        return notas_encontradas
    except FileNotFoundError:
        print(f"ERRO: A pasta '{caminho_pasta}' não foi encontrada.")
        return None
    except Exception as e:
        print(f"ERRO ao escanear a pasta: {e}")
        return None


def gerar_relatorio_excel(df_original, notas_encontradas, notas_faltando, pasta_saida):
    """Cria um novo arquivo Excel com as células coloridas."""
    caminho_relatorio = os.path.join(pasta_saida, "RELATORIO_VERIFICACAO.xlsx")
    print(f"\nGerando relatório colorido em Excel: {caminho_relatorio}")

    # Define os preenchimentos de cor
    fill_verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Cria um novo Workbook do openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório de Verificação"

    # Adiciona o cabeçalho
    ws.append(list(df_original.columns))

    # Itera sobre as linhas do DataFrame original para adicionar os dados e as cores
    for index, row in df_original.iterrows():
        # Converte a linha para uma lista de valores
        valores_linha = list(row.values)
        ws.append(valores_linha)

        # Pega o número do documento da linha atual
        num_doc = str(row['DOCUMENTO']) if pd.notna(row['DOCUMENTO']) else ''

        # Encontra a célula correspondente na planilha
        celula_documento = ws.cell(row=ws.max_row, column=list(df_original.columns).index('DOCUMENTO') + 1)

        # Pinta a célula de acordo com o resultado
        if num_doc in notas_encontradas:
            celula_documento.fill = fill_verde
        elif num_doc:  # Só pinta de vermelho se não estiver vazia
            celula_documento.fill = fill_vermelho

    try:
        wb.save(caminho_relatorio)
        print("Relatório em Excel gerado com sucesso!")
    except Exception as e:
        print(f"ERRO ao salvar o relatório em Excel: {e}")
        print("Verifique se o arquivo não está aberto em outro programa.")


def main():
    """Função principal que orquestra a verificação."""
    print("--- Ferramenta de Verificação de Notas Fiscais em PDF ---")

    caminho_planilha = selecionar_arquivo("Selecione a PLANILHA (Excel ou CSV) com a lista de notas")
    if not caminho_planilha:
        print("Nenhuma planilha selecionada. Encerrando.")
        return

    pasta_pdfs = selecionar_pasta("Selecione a PASTA onde os PDFs foram salvos")
    if not pasta_pdfs:
        print("Nenhuma pasta selecionada. Encerrando.")
        return

    df_original, notas_da_planilha = carregar_notas_da_planilha(caminho_planilha)
    if df_original is None:
        return

    notas_na_pasta = escanear_pasta_pdf(pasta_pdfs)
    if notas_na_pasta is None:
        return

    # Compara os dois conjuntos para encontrar as diferenças
    notas_encontradas = notas_da_planilha.intersection(notas_na_pasta)
    notas_faltando = notas_da_planilha.difference(notas_na_pasta)

    # --- EXIBE O RELATÓRIO NO TERMINAL ---
    print("\n\n--- RELATÓRIO DE VERIFICAÇÃO ---")
    print(f"Total de notas únicas na planilha: {len(notas_da_planilha)}")
    print(f"Total de PDFs encontrados na pasta: {len(notas_na_pasta)}")
    print("------------------------------------")
    print(f"Notas encontradas (planilha E pasta): {len(notas_encontradas)}")
    print(f"Notas FALTANDO (na planilha, mas não na pasta): {len(notas_faltando)}")

    if notas_faltando:
        print("\n>>> LISTA DE NOTAS FISCAIS NÃO ENCONTRADAS <<<")
        # Imprime a lista ordenada para facilitar a visualização
        for nota in sorted(list(notas_faltando)):
            print(f"- {nota}")
    else:
        print("\nÓtima notícia! Todas as notas fiscais da planilha foram encontradas na pasta.")

    # Gera o relatório final em Excel
    gerar_relatorio_excel(df_original, notas_encontradas, notas_faltando, os.path.dirname(caminho_planilha))


if __name__ == "__main__":
    main()
